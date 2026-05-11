"""
Build performance-data.xlsx from performance-data.json
History years are derived dynamically — one Pass/Fail column per year, newest first.
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = Path(r"C:\Users\alpes\GitHub\apra-intelligence-platform\pipeline-data\performance-data.json")
OUT = Path(r"C:\Users\alpes\GitHub\apra-intelligence-platform\pipeline-data\performance-data.xlsx")

FONT_NAME = "Arial"

RAG_FILLS = {
    "Green":   PatternFill("solid", fgColor="FF00B050"),
    "Amber":   PatternFill("solid", fgColor="FFFFC000"),
    "Red":     PatternFill("solid", fgColor="FFFF0000"),
    "Unknown": PatternFill("solid", fgColor="FFD3D3D3"),
}

PF_FILLS = {
    "Pass":    PatternFill("solid", fgColor="FF00B050"),
    "Fail":    PatternFill("solid", fgColor="FFFF0000"),
    "Unknown": PatternFill("solid", fgColor="FFD3D3D3"),
}

HEADER_FILL  = PatternFill("solid", fgColor="FF1F4E79")
HEADER_FONT  = Font(name=FONT_NAME, bold=True, color="FFFFFFFF", size=10)
TITLE_FONT   = Font(name=FONT_NAME, bold=True, size=12, color="FF1F4E79")
NORMAL_FONT  = Font(name=FONT_NAME, size=10)
BOLD_FONT    = Font(name=FONT_NAME, bold=True, size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center")

THIN   = Side(style="thin", color="FFB0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT_FMT = "0.00%"


def all_history_years(records: list[dict]) -> list[str]:
    """Return all unique history years across all records, newest first."""
    years = set()
    for r in records:
        years.update(r.get("history", {}).keys())
    return sorted(years, reverse=True)


def style_header_row(ws, row: int, col_count: int):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border    = BORDER


def style_data_cell(cell, fmt=None, rag=None, pf=None, center=False):
    cell.font      = NORMAL_FONT
    cell.border    = BORDER
    cell.alignment = CENTER_ALIGN if center else LEFT_ALIGN
    if fmt:
        cell.number_format = fmt
    if rag and rag in RAG_FILLS:
        cell.fill = RAG_FILLS[rag]
        cell.font = Font(name=FONT_NAME, size=10, bold=True)
    if pf and pf in PF_FILLS:
        cell.fill = PF_FILLS[pf]
        cell.font = Font(name=FONT_NAME, size=10, bold=True)


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ── Summary ───────────────────────────────────────────────────────────────────
def build_summary(wb, data):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24

    ws["A1"] = "APRA Performance Data — Summary"
    ws["A1"].font = TITLE_FONT

    rows = [
        ("Last Updated",           data["last_updated"]),
        ("MySuper Source Years",   ", ".join(data["source_years_mysuper"])),
        ("TDP Source Years",       ", ".join(data["source_years_tdp"])),
        ("Total MySuper Products", data["total_mysuper_products"]),
        ("Total TDP Options",      data["total_tdp_options"]),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = BOLD_FONT
        ws.cell(row=i, column=1).alignment = LEFT_ALIGN
        ws.cell(row=i, column=2, value=value).font = NORMAL_FONT
        ws.cell(row=i, column=2).alignment = LEFT_ALIGN


# ── MySuper ───────────────────────────────────────────────────────────────────
def build_mysuper(wb, products):
    ws = wb.create_sheet("MySuper Products")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36

    years = all_history_years(products)  # e.g. ["2025", "2024", "2023", ...]

    fixed_headers = [
        "Product Name",
        "Pass / Fail (Current)",
        "10yr NIR p.a.",
        "NIR RAG",
        "Admin Fees ($50k)",
        "Fees $50k RAG",
        "Admin Fees ($100k)",
        "Fees $100k RAG",
    ]
    history_headers = [f"Pass / Fail ({yr})" for yr in years]
    headers = fixed_headers + history_headers

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    for r, p in enumerate(products, start=2):
        pf   = p.get("pass_fail_current") or "Unknown"
        hist = p.get("history", {})

        fixed_values = [
            (p["product_name"],           None,    None,                   None),
            (pf,                          None,    None,                   pf),
            (p.get("nir_10yr"),           PCT_FMT, p.get("nir_rag"),       None),
            (p.get("nir_rag", ""),        None,    p.get("nir_rag"),       None),
            (p.get("fees_50k"),           PCT_FMT, p.get("fees_50k_rag"),  None),
            (p.get("fees_50k_rag", ""),   None,    p.get("fees_50k_rag"),  None),
            (p.get("fees_100k"),          PCT_FMT, p.get("fees_100k_rag"), None),
            (p.get("fees_100k_rag", ""),  None,    p.get("fees_100k_rag"), None),
        ]
        history_values = [
            (hist.get(yr, ""), None, None, hist.get(yr) or "Unknown")
            for yr in years
        ]

        for c, (val, fmt, rag, pf_val) in enumerate(fixed_values + history_values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data_cell(cell, fmt=fmt, rag=rag, pf=pf_val, center=(c != 1))

    # Fixed column widths
    fixed_widths = {"A": 42, "B": 20, "C": 16, "D": 12, "E": 18, "F": 12, "G": 18, "H": 12}
    set_col_widths(ws, fixed_widths)
    # Dynamic history column widths
    for i, yr in enumerate(years):
        col = get_column_letter(len(fixed_headers) + i + 1)
        ws.column_dimensions[col].width = 16


# ── TDP ───────────────────────────────────────────────────────────────────────
def build_tdp(wb, options):
    ws = wb.create_sheet("TDP Options")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36

    years = all_history_years(options)

    fixed_headers = [
        "Product Name",
        "Investment Menu",
        "Investment Option",
        "Product Type",
        "Pass / Fail (Current)",
        "10yr NIR p.a.",
        "NIR RAG",
        "Admin Fees ($50k)",
        "Fees $50k RAG",
        "Admin Fees ($100k)",
        "Fees $100k RAG",
    ]
    history_headers = [f"Pass / Fail ({yr})" for yr in years]
    headers = fixed_headers + history_headers

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    left_cols = set(range(1, 5))  # Product Name, Menu, Option, Type — left-aligned

    for r, o in enumerate(options, start=2):
        pf   = o.get("pass_fail_current") or "Unknown"
        hist = o.get("history", {})

        fixed_values = [
            (o.get("product_name", ""),          None,    None,                    None),
            (o.get("investment_menu_name", ""),   None,    None,                    None),
            (o.get("investment_option_name", ""), None,    None,                    None),
            (o.get("product_type", ""),           None,    None,                    None),
            (pf,                                  None,    None,                    pf),
            (o.get("nir_10yr"),                   PCT_FMT, o.get("nir_rag"),        None),
            (o.get("nir_rag", ""),                None,    o.get("nir_rag"),        None),
            (o.get("fees_50k"),                   PCT_FMT, o.get("fees_50k_rag"),   None),
            (o.get("fees_50k_rag", ""),           None,    o.get("fees_50k_rag"),   None),
            (o.get("fees_100k"),                  PCT_FMT, o.get("fees_100k_rag"),  None),
            (o.get("fees_100k_rag", ""),          None,    o.get("fees_100k_rag"),  None),
        ]
        history_values = [
            (hist.get(yr, ""), None, None, hist.get(yr) or "Unknown")
            for yr in years
        ]

        for c, (val, fmt, rag, pf_val) in enumerate(fixed_values + history_values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data_cell(cell, fmt=fmt, rag=rag, pf=pf_val, center=(c not in left_cols))

    fixed_widths = {
        "A": 36, "B": 36, "C": 36, "D": 18,
        "E": 20, "F": 14, "G": 12,
        "H": 18, "I": 12, "J": 18, "K": 12,
    }
    set_col_widths(ws, fixed_widths)
    for i, yr in enumerate(years):
        col = get_column_letter(len(fixed_headers) + i + 1)
        ws.column_dimensions[col].width = 16


# ── Main ──────────────────────────────────────────────────────────────────────
data = json.loads(SRC.read_text(encoding="utf-8"))
wb   = Workbook()

build_summary(wb, data)
build_mysuper(wb, data["mysuper_products"])
build_tdp(wb, data["tdp_products"])

wb.save(OUT)

ms_years = all_history_years(data["mysuper_products"])
tdp_years = all_history_years(data["tdp_products"])
print(f"Saved: {OUT}")
print(f"  MySuper rows        : {len(data['mysuper_products'])}")
print(f"  MySuper history yrs : {ms_years}")
print(f"  TDP rows            : {len(data['tdp_products'])}")
print(f"  TDP history years   : {tdp_years}")
