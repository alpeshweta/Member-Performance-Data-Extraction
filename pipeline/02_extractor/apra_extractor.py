"""
APRA Excel Extractor
====================
Extracts raw data from APRA CPPP xlsx files using openpyxl.
Part of the APRA Performance Data Loader pipeline — Step 2.

Usage:
    python3 apra_extractor.py --manifest ./apra_data/apra_manifest.json --out ./apra_work
"""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

# ── Column definitions ────────────────────────────────────────────────────────

MYSUPER_COLS_ALL = [
    "RSE licensee MySuper product name",
    "Pass/Fail Indicator",
]

MYSUPER_COLS_CURRENT = [
    "RSE licensee MySuper product name",
    "Pass/Fail Indicator",
    "10 year Net Investment Return (NIR) p.a.",
    "Administration fees and costs charged ($50,000 account balance)",
    "Administration fees and costs charged ($100,000 account balance)",
]

MYSUPER_RAG_COLS = [
    "10 year Net Investment Return (NIR) p.a.",
    "Administration fees and costs charged ($50,000 account balance)",
    "Administration fees and costs charged ($100,000 account balance)",
]

TDP_COLS_ALL = [
    "Product Name",
    "Investment Menu Name",
    "Investment Option Name",
    "Pass/Fail Indicator",
]

TDP_COLS_CURRENT = [
    "Product Name",
    "Investment Menu Name",
    "Investment Option Name",
    "Pass/Fail Indicator",
    "10 year Net Investment Return (NIR) p.a.",
    "Administration fees and costs charged ($50,000 account balance)",
    "Administration fees and costs charged ($100,000 account balance)",
]

TDP_RAG_COLS = [
    "10 year Net Investment Return (NIR) p.a.",
    "Administration fees and costs charged ($50,000 account balance)",
    "Administration fees and costs charged ($100,000 account balance)",
]

# ── Sheet name fallback lists (newest → oldest naming convention) ─────────────
# The extractor tries each name in order and uses the first one found.

MYSUPER_SHEET_NAMES = [
    "MySuper Products",          # 2025
    "MySuper results",           # 2023–2024 combined files
    "Performance test results",  # 2021–2022
]

TDP_NONPLATFORM_SHEET_NAMES = [
    "Non-Platform TDPs",         # 2025
    "Non-Platform TDP results",  # 2023–2024 combined files
]

TDP_PLATFORM_SHEET_NAMES = [
    "Platform TDPs",             # 2025
    "Platform TDP results",      # 2023–2024 combined files
]

# ── Column name aliases (handles APRA year-on-year renames) ──────────────────
# Maps the actual column name found in the file → the canonical name used in
# the COLS constants above.  Add new entries here whenever APRA renames a column.

COLUMN_ALIASES = {
    # 2021–2025 MySuper: product name (all historical variants map to canonical)
    "MySuper product name": "RSE licensee MySuper product name",
    # 2025 (both files): capitalisation change on Pass/Fail
    "Pass/Fail indicator": "Pass/Fail Indicator",
    # 2025 TDP: all three identifier columns lowercased
    "Product name": "Product Name",
    "Investment menu name": "Investment Menu Name",
    "Investment option name": "Investment Option Name",
}


# ── Colour utilities ──────────────────────────────────────────────────────────

def extract_rgb(cell) -> str:
    """Extract a 6-character hex RGB string from a cell's fill colour."""
    try:
        raw = cell.fill.fgColor.rgb  # Returns 8-char AARRGGBB or '00000000'
        if raw and raw != "00000000":
            return raw[-6:].upper()
    except Exception:
        pass
    return ""


def read_colour_legend(wb: openpyxl.Workbook) -> dict[str, str]:
    """
    Read the Colour Legend sheet and return {RGB_hex: label} mapping.
    Expected labels: Green, Amber, Red.
    """
    legend = {}
    if "Colour Legend" not in wb.sheetnames:
        return legend

    ws = wb["Colour Legend"]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                label = cell.value.strip()
                if label in ("Green", "Amber", "Red"):
                    rgb = extract_rgb(cell)
                    if rgb:
                        legend[rgb] = label

    # Fallback: check adjacent cells — sometimes the colour is in the next cell
    if not legend:
        for row in ws.iter_rows():
            cells = list(row)
            for i, cell in enumerate(cells):
                if cell.value and isinstance(cell.value, str):
                    label = cell.value.strip()
                    if label in ("Green", "Amber", "Red"):
                        for offset in [-1, 1]:
                            idx = i + offset
                            if 0 <= idx < len(cells):
                                rgb = extract_rgb(cells[idx])
                                if rgb:
                                    legend[rgb] = label
    return legend


# ── Column index resolution ───────────────────────────────────────────────────

def _normalise_header(s: str) -> str:
    """Collapse internal newlines and excess whitespace in a header string."""
    return re.sub(r"\s+", " ", s.replace("\n", " ")).strip()


def build_col_index(ws) -> tuple[dict[str, int], int]:
    """Return ({column_header: 0-based-col-index}, header_row_1indexed).

    Headers are normalised (internal newlines → space) and known APRA
    year-on-year renames are resolved via COLUMN_ALIASES.

    The row is accepted only when it contains at least one recognised column
    name, preventing group-heading rows from being mistaken for the header row.
    Returning the header row number lets callers start data iteration at
    header_row + 1, skipping the header itself cleanly.
    """
    known = set(COLUMN_ALIASES.keys()) | set(COLUMN_ALIASES.values())
    known |= set(MYSUPER_COLS_ALL) | set(MYSUPER_COLS_CURRENT)
    known |= set(TDP_COLS_ALL) | set(TDP_COLS_CURRENT)

    best_headers: dict[str, int] = {}
    best_row_num = 1

    for row in ws.iter_rows(min_row=1, max_row=15):
        row_num = row[0].row if row else 0
        headers: dict[str, int] = {}
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                key = _normalise_header(cell.value)
                headers[key] = cell.column - 1  # 0-based
        if len(headers) < 3:
            continue
        # Apply aliases before checking membership.
        for alias, canonical in COLUMN_ALIASES.items():
            if alias in headers and canonical not in headers:
                headers[canonical] = headers[alias]
        # Accept this row only if it contains at least one recognised column.
        if headers.keys() & known:
            return headers, row_num
        # Keep the widest candidate as a last-resort fallback.
        if len(headers) > len(best_headers):
            best_headers = headers
            best_row_num = row_num

    return best_headers, best_row_num


def _find_sheet(wb: openpyxl.Workbook, candidates: list[str]) -> str | None:
    """Return the first sheet name from candidates that exists in the workbook."""
    for name in candidates:
        if name in wb.sheetnames:
            return name
    return None


# ── Extraction helpers ────────────────────────────────────────────────────────

def extract_mysuper(ws, is_current_year: bool, colour_legend: dict,
                    source_year: int, source_file: str) -> list[dict]:
    col_index, header_row = build_col_index(ws)
    required = MYSUPER_COLS_CURRENT if is_current_year else MYSUPER_COLS_ALL
    missing = [c for c in required if c not in col_index]
    if missing:
        print(f"  [warn] MySuper — missing columns: {missing}")

    id_col = col_index.get("RSE licensee MySuper product name")
    if id_col is None:
        print(f"  [warn] MySuper product name column not found in {source_file}")
        return []

    records = []
    data_started = False

    # Start iteration from the row after the header to avoid capturing header text as data.
    for row in ws.iter_rows(min_row=header_row + 1):
        cells = list(row)
        if not cells or id_col >= len(cells):
            continue

        product_name = cells[id_col].value
        if not isinstance(product_name, str) or not product_name.strip():
            continue

        data_started = True
        rec = {
            "source_year": source_year,
            "product_type": "MySuper",
            "is_current_year": is_current_year,
            "source_file": source_file,
            "product_name": product_name.strip(),
        }

        pf_col = col_index.get("Pass/Fail Indicator")
        if pf_col is not None and pf_col < len(cells):
            rec["pass_fail_raw"] = str(cells[pf_col].value or "").strip()

        if is_current_year:
            for col_name in MYSUPER_RAG_COLS:
                col_idx = col_index.get(col_name)
                if col_idx is not None and col_idx < len(cells):
                    cell = cells[col_idx]
                    rec[col_name] = cell.value
                    rgb = extract_rgb(cell)
                    rag = colour_legend.get(rgb, "Unknown") if rgb else "Unknown"
                    rec[col_name + "_rag_rgb"] = rgb
                    rec[col_name + "_rag"] = rag

        records.append(rec)

    if not data_started:
        print(f"  [warn] MySuper sheet — no data rows found in {source_file}")
    return records


def extract_tdp(ws, sheet_name: str, is_current_year: bool, colour_legend: dict,
                source_year: int, source_file: str) -> list[dict]:
    col_index, header_row = build_col_index(ws)
    required = TDP_COLS_CURRENT if is_current_year else TDP_COLS_ALL
    missing = [c for c in required if c not in col_index]
    if missing:
        print(f"  [warn] {sheet_name} — missing columns: {missing}")

    product_type = "Platform TDP" if "Platform" in sheet_name else "Non-Platform TDP"

    # Historical TDP files (2023–2024) lack a Product Name column.
    # Fall back to Investment Option Name as the primary row identifier.
    prod_col = col_index.get("Product Name")
    using_option_as_product = False
    if prod_col is None:
        prod_col = col_index.get("Investment Option Name")
        using_option_as_product = True
        if prod_col is not None:
            print(f"  [info] {sheet_name}: no 'Product Name' column — using 'Investment Option Name' as product identifier.")

    if prod_col is None:
        print(f"  [warn] {sheet_name}: no usable row-identifier column found. Skipping.")
        return []

    records = []

    for row in ws.iter_rows(min_row=header_row + 1):
        cells = list(row)
        if not cells or prod_col >= len(cells):
            continue

        product_name = cells[prod_col].value
        if not isinstance(product_name, str) or not product_name.strip():
            continue

        rec = {
            "source_year": source_year,
            "product_type": product_type,
            "is_current_year": is_current_year,
            "source_file": source_file,
            "product_name": product_name.strip(),
        }

        if not using_option_as_product:
            # Standard 2025 columns
            menu_col = col_index.get("Investment Menu Name")
            if menu_col is not None and menu_col < len(cells):
                rec["investment_menu_name"] = str(cells[menu_col].value or "").strip()

        # Investment Option Name
        opt_col = col_index.get("Investment Option Name")
        if opt_col is not None and opt_col < len(cells):
            rec["investment_option_name"] = str(cells[opt_col].value or "").strip()
        elif using_option_as_product:
            # product_name IS the option name — copy it over
            rec["investment_option_name"] = product_name.strip()

        pf_col = col_index.get("Pass/Fail Indicator")
        if pf_col is not None and pf_col < len(cells):
            rec["pass_fail_raw"] = str(cells[pf_col].value or "").strip()

        if is_current_year:
            for col_name in TDP_RAG_COLS:
                col_idx = col_index.get(col_name)
                if col_idx is not None and col_idx < len(cells):
                    cell = cells[col_idx]
                    rec[col_name] = cell.value
                    rgb = extract_rgb(cell)
                    rag = colour_legend.get(rgb, "Unknown") if rgb else "Unknown"
                    rec[col_name + "_rag_rgb"] = rgb
                    rec[col_name + "_rag"] = rag

        records.append(rec)

    return records


# ── File classifier ───────────────────────────────────────────────────────────

def classify_file(entry: dict) -> tuple[str, int, bool]:
    """Return (file_type, year, is_current_year) from a manifest entry."""
    product_type = entry.get("product_type", "Unknown")
    year = int(entry.get("year", 0))
    is_current = bool(entry.get("is_current_year", False))
    return product_type, year, is_current


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="APRA Excel data extractor")
    parser.add_argument("--manifest", required=True, help="Path to apra_manifest.json")
    parser.add_argument("--out", default="./apra_work", help="Output directory")
    args = parser.parse_args()

    manifest_path = Path(args.manifest)
    output_dir = Path(args.out)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not manifest_path.exists():
        print(f"ERROR: Manifest not found: {manifest_path}", file=sys.stderr)
        return 1

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    downloaded = [e for e in manifest if e.get("status") == "downloaded"]

    print(f"[setup] Processing {len(downloaded)} downloaded files.")

    all_records = []
    combined_legend = {}
    extraction_log = []

    for entry in downloaded:
        file_path = Path(entry.get("file_path", ""))
        if not file_path.exists():
            print(f"[skip] File not found: {file_path}")
            extraction_log.append({"file": str(file_path), "status": "skipped", "reason": "file not found"})
            continue

        product_type, year, is_current_year = classify_file(entry)
        print(f"\n[extract] {file_path.name} — {product_type} {year} (current={is_current_year})")

        try:
            wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        except Exception as exc:
            print(f"  [error] Cannot open file: {exc}")
            extraction_log.append({"file": file_path.name, "status": "failed", "reason": str(exc)})
            continue

        # Read colour legend from current year files
        legend = {}
        if is_current_year:
            legend = read_colour_legend(wb)
            if legend:
                combined_legend.update(legend)
                print(f"  [legend] Colour map: {legend}")
            else:
                print(f"  [warn] No colour legend found in {file_path.name}. RAG will be derived from Pass/Fail.")

        records = []
        sheets_processed = []

        if product_type == "MySuper":
            sheet_name = _find_sheet(wb, MYSUPER_SHEET_NAMES)
            if sheet_name:
                ws = wb[sheet_name]
                recs = extract_mysuper(ws, is_current_year, legend, year, file_path.name)
                records.extend(recs)
                sheets_processed.append(f"{sheet_name} ({len(recs)} records)")
            else:
                tried = ", ".join(f"'{s}'" for s in MYSUPER_SHEET_NAMES)
                print(f"  [warn] No MySuper sheet found in {file_path.name} (tried {tried})")
                sheets_processed.append("MySuper — NOT FOUND")

        elif product_type == "TDP":
            for candidates, label in [
                (TDP_NONPLATFORM_SHEET_NAMES, "Non-Platform TDP"),
                (TDP_PLATFORM_SHEET_NAMES,    "Platform TDP"),
            ]:
                sheet_name = _find_sheet(wb, candidates)
                if sheet_name:
                    ws = wb[sheet_name]
                    recs = extract_tdp(ws, sheet_name, is_current_year, legend, year, file_path.name)
                    records.extend(recs)
                    sheets_processed.append(f"{sheet_name} ({len(recs)} records)")
                else:
                    tried = ", ".join(f"'{s}'" for s in candidates)
                    print(f"  [warn] No {label} sheet found in {file_path.name} (tried {tried})")
                    sheets_processed.append(f"{label} — NOT FOUND")

        wb.close()
        all_records.extend(records)
        extraction_log.append({
            "file": file_path.name,
            "year": year,
            "product_type": product_type,
            "is_current_year": is_current_year,
            "records": len(records),
            "sheets": sheets_processed,
            "status": "extracted",
        })
        print(f"  [ok] {len(records)} records extracted.")

    # Check if any current year RAG was all Unknown
    current_rag_records = [r for r in all_records if r.get("is_current_year")]
    rag_all_unknown = all(
        r.get("10 year Net Investment Return (NIR) p.a._rag", "Unknown") == "Unknown"
        for r in current_rag_records
    ) if current_rag_records else False

    if rag_all_unknown:
        print("\n[WARN] All RAG values are 'Unknown' for current year records.")
        print("       APRA may be using conditional formatting — openpyxl cannot read it.")
        print("       RAG status will be derived from Pass/Fail in the Data Cleaner step.")

    # Save outputs
    extracted_path = output_dir / "apra_extracted.json"
    legend_path    = output_dir / "apra_colour_legend.json"
    log_path       = output_dir / "apra_extraction_log.json"

    extracted_path.write_text(json.dumps(all_records, indent=2, default=str), encoding="utf-8")
    legend_path.write_text(json.dumps(combined_legend, indent=2), encoding="utf-8")
    log_path.write_text(json.dumps(extraction_log, indent=2, default=str), encoding="utf-8")

    print(f"\n{'=' * 50}")
    print(f"APRA Excel Extractor — Complete")
    print(f"  Total records      : {len(all_records)}")
    print(f"  Colour legend keys : {len(combined_legend)}")
    print(f"  RAG all unknown    : {rag_all_unknown}")
    print(f"  Extracted JSON     : {extracted_path}")
    print(f"  Colour legend      : {legend_path}")
    print(f"  Extraction log     : {log_path}")
    print(f"{'=' * 50}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
